# -*- coding: utf-8 -*-
"""
Controller.py
@author: Groß, Hendrik
"""
import os, asyncio
import pathlib, importlib,inspect
import openpyxl
import pandapower as pp

"""
Controller definition for the C-HiL (Controller-Hardware-in-the-Loop) implementation
"""
from .IModbusElement import IModbusElement
from pymodbus.client import AsyncModbusTcpClient
from .Enum import ReadMode

class Controller():
    """
    Controller for the C-HiL (Controller-Hardware-in-the-Loop) process using a real-time simulation on an OPAL-RT 4510.
    Manages communication between the pandapower network model and the OPAL-RT 4510 real-time simulator (Modbus slave). The implementation can
    either actively interact with the real-time simulation (control it) or be used purely to visualize the results.
    
    **Functionality:**
        - Asynchronous TCP/IP Modbus communication
        - Management of network components (bus, load, generator, switch)
        - Bidirectional synchronization between pandapower and the HIL simulator
        - Configuration management (load/save from/to Excel)
        - Persistent event loop for multiple asynchronous operations
    """

    # Constructor
    def __init__(self,_host:str = "localhost", _port:int = 502, _pp_net:pp.pandapowerNet | None = None, _component_list:list[IModbusElement] = []):
        """
        Initializes the controller.

        :param _host: IP address of the OPAL-RT 4510 (e.g. "10.24.3.38")
        :type _host: str
        :param _port: Modbus TCP port of the OPAL-RT 4510 (default: 1502)
        :type _port: int
        :param _pp_net: PandaPower network model for communication
        :type _pp_net: pp.pandapowerNet | None
        :param _component_list: List of controllable and readable network components (bus, load, generator, etc.)
        :type _component_list: list[IModbusElement] | None
        """
        # OPAL-RT connection information
        self.host: str = _host                                          # Modbus slave host (e.g. 'localhost') 
        self.port: int = _port                                          # Modbus slave port (e.g. '502')

        # Asynchronous Modbus communication  
        self.event_loop: asyncio.AbstractEventLoop | None = None        # Persistent event loop for all async operations
        self.modbus_client: AsyncModbusTcpClient | None = None          # Modbus master (TCP/IP client)

        # Controllable network components
        self.component_list: list[IModbusElement] = _component_list     # List of Modbus components that can have multiple Modbus values

        # Set the PandaPower network model
        self.pp_net: pp.pandapowerNet | None = _pp_net                 
        if self.pp_net is not None:
            self.set_pp_net(self.pp_net)

    def set_pp_net(self, _pp_net:pp.pandapowerNet):
        """
        Sets the pandapower network model and distributes it to all components.
        """
        self.pp_net= _pp_net 

        # Transfer the network model to the components when available
        for component in self.component_list:
            component.pp_net = self.pp_net

    def connect(self,_host:str | None = None, _port:int | None = None):
        """
        Establishes a persistent Modbus connection synchronously.

        The method internally uses ``__connect_async`` via the persistent event loop.
        
        :param _host: Optional IP address (overrides the __init__ value if provided)
        :type _host: str
        :param _port: Optional Modbus port (overrides the __init__ value if provided)
        :type _port: int
        :return: The connected AsyncModbusTcpClient or a ConnectionError
        :rtype: AsyncModbusTcpClient
        :raises ConnectionError: If the connection fails or no components are available
        """
        # Create a persistent event loop for all async operations if it does not already exist via connect
        self.__check_event_loop()

        # Use the persistent event loop instead of asyncio.run()
        return self.event_loop.run_until_complete(self.__connect_async(_host,_port)) # pyright: ignore[reportOptionalMemberAccess] -> __check_event_loop

    def test_modbus(self,_host:str | None = None, _port:int | None = None) -> bool:
        """
        Runs a diagnostic test of the Modbus communication.

        A temporary connection is established and then it is checked whether all
        configured components can be read without errors.
        
        :param _host: Optional IP address of the OPAL-RT 4510
        :type _host: str
        :param _port: Optional Modbus TCP port
        :type _port: int
        :return: True if the test is successful, False on error
        :rtype: bool
        
        """
        # Create a persistent event loop for the test once:
        event_loop = asyncio.new_event_loop()
        asyncio.set_event_loop(event_loop)

        # Use the persistent event loop instead of asyncio.run()
        return event_loop.run_until_complete(self.__test_modbus_async(_host,_port))

    def read_all(self, mode: ReadMode = ReadMode.ALL):
        """
        Reads all components synchronously and in parallel.

        Internally, the asynchronous read access is bundled via ``asyncio.gather``.
        
        :return: Result of the asynchronous operations
        :raises ConnectionError: If no active Modbus connection is available
        """
        # It is possible that no persistent event loop has been set yet if there is no Modbus connection
        self.__check_event_loop()

        try:
            return self.event_loop.run_until_complete(self.__read_all_async(mode)) # pyright: ignore[reportOptionalMemberAccess] -> __check_event_loop
        except Exception as e:
            raise ValueError(f"Error 'read_all()': {e}")

    def disconnect(self):
        """
        Disconnects the Modbus connection and cleans up component references.

        The Modbus client is removed from all components and the
        active connection is closed.
        """
        # Remove the Modbus client from the components
        if self.component_list:
            for component in self.component_list:
                component.set_modbus_client(None)

        # Close the connection
        if self.modbus_client and self.modbus_client.connected:
            self.modbus_client.close()

    # --------------------------------------------------------------------------------------------------------
    # Internal methods (asynchronous):
    def __check_event_loop(self):
        """Ensures that a persistent event loop exists."""
        if self.event_loop is None or self.event_loop.is_closed():
            self.event_loop = asyncio.new_event_loop()
            # Pass the event loop to all network elements
            for cb in self.component_list:
                cb.event_loop = self.event_loop
            asyncio.set_event_loop(self.event_loop)

    async def __test_modbus_async(self,_host:str | None = None, _port:int | None = None) -> bool:
        """
        Asynchronous diagnostic test for connection and register access.

        Establishes a temporary connection, reads all components for testing,
        and then closes the connection again.
        
        :param _host: IP of the OPAL-RT 4510
        :type _host: str
        :param _port: Port of the OPAL-RT 4510
        :type _port: int
        :return: True on successful test, False on error
        :rtype: bool
        """
        if _host is None:
            _host = self.host
        if _port is None:
            _port = self.port
        
        print(f"CONNECTION_CHECK - Modbus-Slave: Host='{self.host}'; Port='{self.port}'")
        try:
            # Establish a connection for testing and close it afterwards
            _modbus_client = AsyncModbusTcpClient(host=_host,port=_port)  

            # Open the test connection
            if not await _modbus_client.connect():
                raise ConnectionError(f"CONNECTION_CHECK_FAILED - Connection failed: {self.host}:{self.port}")
            else:
                print(f"CONNECTION_CHECK_SUCCESS - Connection successfull {self.host}:{self.port}")

            # Check whether all Modbus values can also be read from the Modbus slave 
            print("MODBUS_REG_CHECK - Checking whether all register values of the network components can be read from the Modbus slave...")
            for component in self.component_list:
                component.set_modbus_client(_modbus_client)
                try:
                    if await component.read_async() == False:
                        raise ValueError(f"MODBUS_REG_CHECK_FAILED - Register value could not be read '{component}'")
                except Exception as e:
                    raise ValueError(f"MODBUS_REG_CHECK_FAILED - '{component}' - {e}")
            print("MODBUS_REG_CHECK_SUCCESS -All register values could be read successfully from the Modbus slave")

            # Reset the Modbus client in the components
            for component in self.component_list:
                component.set_modbus_client(None)

            # Close the connection
            if _modbus_client and _modbus_client.connected:
                _modbus_client.close()
        except Exception as e:
            print(e)
            return False
        print("MODBUS_CHECK_SUCCESS - Modbus communication test successful!")
        return True

    async def __read_all_async(self, mode: ReadMode = ReadMode.ALL):
        """
        Reads all components asynchronously in parallel.

        An async task is created for each component and executed together.
        
        :return: Result of asyncio.gather (all task results)
        :raises ConnectionError: If no Modbus connection exists or it is not connected
        """
        # Read all components in parallel
        tasks = []
        for component in self.component_list:
            tasks.append(component.read_async(mode))

        await asyncio.gather(*tasks, return_exceptions=True)

    async def __connect_async(self,_host:str | None = None, _port:int | None = None):
        """
        Establishes a persistent Modbus TCP connection asynchronously.

        Optional host/port values override the current
        controller configuration. On success, the client is distributed to all components
        and stored as an instance attribute.
        
        :param _host: Optional IP address (overrides self.host)
        :type _host: str
        :param _port: Optional port (overrides self.port)
        :type _port: int
        :return: The connected AsyncModbusTcpClient
        :rtype: AsyncModbusTcpClient
        :raises ConnectionError: If connect fails or no components are available
        """
        # Override the OPAL-RT connection information if provided
        if _host is not None:
            self.host = _host
        if _port is not None:
            self.port = _port  
        
        # If there is no list of components, the Modbus slave is not useful either :(
        if self.component_list is None or len(self.component_list) == 0:
            raise ConnectionError(f"NO_GRID_ELEMENTS - No grid elemets available. Connection to {str(self.host)}:{str(self.port)} is terminated!")
        
        # Test the connection and store it as an object if successful
        print(f"CONNECTION_CHECK - Modbus-Slave: Host='{self.host}'; Port='{self.port}'")
        self.modbus_client = None
        _modbus_client = AsyncModbusTcpClient(host=self.host,port=self.port)
        if not await _modbus_client.connect():
            raise ConnectionError(f"CONNECTION_CHECK_FAILED - Connection failed: {self.host}:{self.port}")
        print(f"CONNECTION_CHECK_SUCCESS - Connection successfull {self.host}:{self.port}")
        
        # Pass the Modbus TCP/IP client to the components for reading
        print(f"MODBUS_CLIENT - Passing the Modbus client to the components...")
        _index = 1
        for component in self.component_list:
            print(f"[{_index}|{len(self.component_list)}] - {component}")
            component.set_modbus_client(_modbus_client)
            _index += 1

        self.modbus_client = _modbus_client
        return _modbus_client

    # --------------------------------------------------------------------------------------------------------
    # STATIC METHODS
    @staticmethod
    def get_class_map() -> dict[str, type]:
        """
        Automatically determines all available component classes.

        All Python files in the package directory are scanned and classes
        that inherit from ``IModbusElement`` are collected.
        
        :return: Dictionary with class names as keys and class objects as values
        :rtype: dict[str, type]
        """
        class_map = {}
        # Consider all object classes in the current folder that are subclasses of IModbusElement
        current_py = pathlib.Path(__file__)
        for py_file in current_py.parent.glob("*.py"):
            # Skip the current file
            if py_file.resolve() == current_py.resolve():
                continue

            # Import as a package module so relative imports work
            module = importlib.import_module(f"{__package__}.{py_file.stem}")
            # Find all classes in the module
            for name, cls in inspect.getmembers(module, inspect.isclass):
                if issubclass(cls, IModbusElement) and cls is not IModbusElement:
                    class_map[name] = cls
        # Return the dictionary
        return class_map

    @staticmethod
    def load_cfg_from_excel(_excel_file: str) -> list[IModbusElement]:
        """
        Loads components and Modbus values from an Excel configuration.

        Expects the ``Components`` and ``ModbusValues`` worksheets and creates
        the corresponding component instances from them.
        
        :param _excel_file: Path to the Excel configuration file (.xlsx)
        :type _excel_file: str
        :return: List of loaded component instances
        :rtype: list[IModbusElement]
        :raises KeyError: If a component class is not found
        :raises Exception: If the Excel format is invalid or worksheets are missing
        """
        # Import the base classes for the __init__ function
        from .Enum import ModbusDataType, DataType
        from .Value import ModbusValue

        # Read possible subclasses for import
        class_map = Controller.get_class_map()
       
        # Load the list of objects from the Excel file
        component_list = []

        # Load the Excel file
        wb = openpyxl.load_workbook(_excel_file)
        ws_components = wb["Components"]
        ws_regvals = wb["ModbusValues"]

        # Iterate over all components
        for component_row in ws_components.iter_rows():
            # Skip the first row (headers)
            if component_row[0].row > 1:    # pyright: ignore[reportOptionalOperand] 
                _uuid = component_row[0].value
                
                # Read Modbus values
                values = {}
                for regval_row in ws_regvals.iter_rows():
                    # Find Modbus values for the corresponding object
                    if regval_row[0].value == _uuid:
                        
                        # Create the ModbusValue data object
                        modbus_data_type = ModbusDataType(regval_row[2].value)
                        data_type = DataType.from_label(regval_row[3].value)    # pyright: ignore[reportArgumentType]
                        adr: int = regval_row[4].value                          # pyright: ignore[reportAssignmentType]
                        unit: str | None = regval_row[5].value                  # pyright: ignore[reportAssignmentType]
                        scale: float = regval_row[6].value                      # pyright: ignore[reportAssignmentType]
                        # Create object
                        if unit == "":
                            unit = None
                        # Create object
                        modbus_value = ModbusValue(modbus_data_type,data_type,adr,unit,scale)

                        # Store object in dictionary
                        attribute_name = regval_row[1].value
                        values[attribute_name] = modbus_value

                # Read the component attributes
                class_name: str = component_row[1].value    # pyright: ignore[reportAssignmentType]
                obj_name: str = component_row[2].value      # pyright: ignore[reportAssignmentType]
                net_index: int = component_row[3].value     # pyright: ignore[reportAssignmentType]

                # Create object from class
                if class_name not in class_map:
                    raise KeyError(f"Class intance for '{class_name}' could not be initialized! Class not found in the class map.")
                cls = class_map[class_name]
                component = cls(obj_name,net_index, values)

                # Add object to list
                component_list.append(component)

        return component_list

    @staticmethod
    def write_cfg_to_excel(_component_list:list[IModbusElement], _excel_file: str):
        """
        Saves components and Modbus values to an Excel configuration.

        The ``Components`` and ``ModbusValues`` worksheets are created.
        The file is compatible with ``load_cfg_from_excel``.
        
        :param _component_list: List of components to save
        :type _component_list: list[IModbusElement]
        :param _excel_file: Target file path for the Excel file (.xlsx)
        :type _excel_file: str
        """

         # If no path is provided initially, use the default 
        if _excel_file is None or _excel_file == "":
            _excel_file = os.getcwd() + "\\ExcelConfig.xlsx"
        
        # Create Excel file with worksheets
        wb = openpyxl.Workbook()

        # Worksheet: Components
        if wb.active is None:
            wb.active = wb.create_sheet("Components")
        wb.active.title = "Components"
        wb.active.append(["ID", "Objektklasse","Name", "NetIndex"])

        # Worksheet: ModbusValues
        sheet_rv = wb.create_sheet("ModbusValues")
        sheet_rv.append(["ComponentID", "Attribute", "ModbusDataType", "DataType", "Address","Unit","Scale"])

        # Write all components (each component saves itself + its Modbus values)
        for component in _component_list:
            component.write_to_excel(wb)

        # Save
        wb.save(_excel_file)