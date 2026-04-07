 # -*- coding: utf-8 -*-
"""
PPtoePHASORSIM.py
@author: Groß, Hendrik
"""
from pathlib import Path
import pathlib, importlib,inspect
import os, shutil, uuid
import pandapower as pp
from openpyxl import load_workbook
from dataclasses import dataclass
from typing import List
from enum import Enum

# Determine the module path to load subclasses automatically
MODULE_DIR = pathlib.Path(__file__).parent

# Import the base class for building the dynamic network element registry.
from .Base import BaseComponent

@dataclass
class Pin:
    """Represents an I/O pin definition for the OPAL-RT Excel export.

    A pin maps an incoming or outgoing signal to one or more network element IDs, e.g. for connecting to Simulink/Modbus slave.
    """
    class Type(Enum):
        """Pin direction for bidirectional signal exchange."""
        INCOMING = 0   # Signals entering the network model.
        OUTGOING = 1   # Signals provided by the network model.

    # Data fields
    pin_type: Type              # Pin direction (INCOMING or OUTGOING)
    label: str                  # Pin label (e.g. "PV1_Power", "Bus3_Voltage")
    component_list: List[str]   # Linked network element IDs

    def to_row(self) -> List[str]:
        """Converts this pin into an Excel row.

        Output format: [``type_lowercase``, ``label``, ``component_1``, ``component_2``, ...].
        """
        # Resolve pin type names robustly.
        _type_name = ""
        if isinstance(self.pin_type, Pin.Type):
            _type_name = self.pin_type.name
        elif isinstance(self.pin_type, int):
            _type_name = Pin.Type(self.pin_type).name
        else:
            _type_name = str(self.pin_type)

        # Build the output row: [``type_lowercase``, ``label``, *components].
        _data = [_type_name.lower(), self.label] 
        _data.extend(self.component_list)     
        return _data

class Converter():
    """
    Converts pandapower network models into OPAL-RT/ePHASORSIM Excel templates.

    The generated workbook can be used in the ePHASORSIM solver block with the configuration option "Excel".
    """
    # Component classes considered during conversion. This list is automatically populated with all BaseComponent subclasses.
    COMPONENTS = []
    # Automatically detect available classes. Iterate over all Python files in this package directory.
    current_py = pathlib.Path(__file__)
    for py_file in current_py.parent.glob("*.py"):
        # Skip this file (PPtoePHASORSIM.py) to avoid recursive import loops.
        if py_file.resolve() == current_py.resolve():
            continue
        # Import as a package module so relative imports remain valid.
        module = importlib.import_module(f"{__package__}.{py_file.stem}")
        # Register only found Base subclasses from the module.
        for name, cls in inspect.getmembers(module, inspect.isclass):
            if issubclass(cls, BaseComponent) and cls is not BaseComponent:
                COMPONENTS.append(cls)
    
    # Constructor
    def __init__(self, _pp_net : pp.pandapowerNet | None = None):
        """Initializes with an optional pandapower network model."""
        # pandapower network model
        if _pp_net is not None and not isinstance(_pp_net, pp.pandapowerNet):
            raise TypeError(f"'_pp_net' must be a pandapower network model: {type(element).__name__}")
        self.pp_net : pp.pandapowerNet | None = _pp_net

        # Normalize immediately if a model was provided.
        if self.pp_net is not None:
            self.__convert_pp_net_after_load(self.pp_net)

    def load_pp_net_from_pickle(self, _pickle_file: str):
        """Loads a pandapower network from a pickle file and normalizes it."""
        _pp_net = pp.from_pickle(_pickle_file, True)

        # Use the filename stem if the network has no explicit name.
        if _pp_net.name is None or _pp_net.name == "":
            _pp_net.name = Path(_pickle_file).stem

        # Normalize the loaded network model.
        self.__convert_pp_net_after_load(_pp_net)

    def load_pp_net_from_excel(self, _excel_file: str):
        """Loads a pandapower network from an Excel file and normalizes it."""
        # Load network data from Excel.
        _pp_net = pp.from_excel(_excel_file, True)
        
        # Use the filename stem if the network has no explicit name.
        if _pp_net.name is None or _pp_net.name == "":
            _pp_net.name = Path(_excel_file).stem

        # Normalize the loaded network model.
        self.__convert_pp_net_after_load(_pp_net)

    def load_pp_net_from_sqlite(self, _sqlite_file: str):
        """Loads a pandapower network from SQLite and normalizes it."""
        # Load network data from SQLite.
        _pp_net = pp.from_sqlite(_sqlite_file)

        # Use the filename stem if the network has no explicit name.
        if _pp_net.name is None or _pp_net.name == "":
            _pp_net.name = Path(_sqlite_file).stem

        # Normalize the loaded network model.
        self.__convert_pp_net_after_load(_pp_net)

    def __convert_pp_net_after_load(self, _pp_net: pp.pandapowerNet):
        """Normalizes a loaded pandapower model for the OPAL-RT/ePHASORSIM Excel export.

        Main steps:
        1. Normalize all switches to bus-bus form.
        2. Clean network element IDs.
        3. Apply scaling factors to loads and static generators.
        """
        # ================================================================================================================
        # NETWORK MODEL NORMALIZATION
        # ================================================================================================================
        # 1. SWITCH NORMALIZATION: unify bus-line, bus-transformer, and bus-bus links.
        #    - In pandapower, switch objects can be connected to different element types.
        #    - OPAL-RT requires all switches to be represented as bus-bus links (et="b").
        #    - For non-bus switches, an auxiliary bus is created and the original element is reconnected.
        #    - The original switch is replaced by a new bus-bus switch.
        # ================================================================================================================
        for idx, switch in _pp_net.switch.iterrows():
            # Process only non-bus-bus switches (et="b" is already compliant).
            if switch.et != "b":   
                # Keep the switch properties for the replacement bus-bus switch.
                _state = switch.closed        # Switching state (open/closed)
                _from_bus = switch.bus        # Source bus
                _type = switch.type           # Switch type (CB, DS, LS, etc.)
                _name = switch.name           # Switch name
                _z_ohm = switch.z_ohm         # Switch resistance (optional)
                _in_ka = switch.in_ka         # Rated current (optional)

                # Create an auxiliary bus as an intermediate connection point.
                _to_bus = pp.create_bus(_pp_net, vn_kv=_pp_net.bus.at[switch.bus, "vn_kv"], 
                                        name= f"B{str(uuid.uuid4())}", type="n")            

                # Handle line switches (et="l").
                if switch.et == "l":
                    _line = _pp_net.line.loc[switch.element]
                    # Reconnect the side that currently points to the source bus.
                    if _line.from_bus == _from_bus:
                        _pp_net.line.loc[switch.element, "from_bus"] = _to_bus  # Replace from_bus.
                    else:
                        _pp_net.line.loc[switch.element, "to_bus"] = _to_bus    # Replace to_bus.
                                        
                # Handle transformer switches (et="t").
                elif switch.et == "t":
                    _trafo = _pp_net.trafo.loc[switch.element]
                    # Reconnect the affected transformer side.
                    if _trafo.hv_bus == _from_bus:
                        _pp_net.trafo.loc[switch.element, "hv_bus"] = _to_bus   # Replace HV side.
                    else:
                        _pp_net.trafo.loc[switch.element, "lv_bus"] = _to_bus   # Replace LV side.
                    
                # Remove the original switch from the network.
                _pp_net.switch.drop(idx, inplace=True)
                
                # Create an ePHASORSIM-compatible bus-bus replacement switch.
                pp.create_switch(_pp_net, 
                                 bus=_from_bus,           # Original bus.
                                 element=_to_bus,         # Target: auxiliary bus.
                                 et="b",                  # Element type: bus-to-bus.
                                 closed=_state,           # Preserve switch state.
                                 name = _name,            # Preserve switch name.
                                 type=_type,              # Preserve switch type.
                                 z_ohm=_z_ohm,            # Preserve resistance.
                                 in_ka=_in_ka)            # Preserve rated current.
        # ================================================================================================================
        # 2. Clean IDs.
        # Performed in subclasses so updated IDs can be written back into pandapower tables.
        # ================================================================================================================
        for component in self.COMPONENTS:
            _pp_net = component.clean_and_update_ids(_pp_net)

        # ================================================================================================================
        # 3. Apply scaling in the ``sgen`` and ``load`` tables, then set scaling to 1.
        # ================================================================================================================
        for idx, _sgen in _pp_net.sgen.iterrows():
            _pp_net.sgen.loc[idx,"sn_mva"] = _sgen.sn_mva * _sgen.scaling
            _pp_net.sgen.loc[idx,"p_mw"] = _sgen.p_mw * _sgen.scaling
            _pp_net.sgen.loc[idx,"q_mvar"] = _sgen.q_mvar * _sgen.scaling
            _pp_net.sgen.loc[idx,"scaling"] = 1
        for idx, _load in _pp_net.load.iterrows():
            _pp_net.load.loc[idx,"sn_mva"] = _load.sn_mva * _load.scaling
            _pp_net.load.loc[idx,"p_mw"] = _load.p_mw * _load.scaling
            _pp_net.load.loc[idx,"q_mvar"] = _load.q_mvar * _load.scaling
            _pp_net.load.loc[idx,"scaling"] = 1

        # Run a power flow to verify network consistency after normalization.
        pp.runpp(_pp_net, voltage_depend_loads=False)   # Loads are Constant...       
        
        # Store the normalized network model.
        self.pp_net = _pp_net

    def create_excel_template(self, _excel_file: str, _list_pins: List[Pin] | None = None):
        """Creates the OPAL-RT/ePHASORSIM Excel template from the current network.

        The workbook contains general network settings, optional I/O pin mappings, and one worksheet per network element type.
        """
        if _list_pins is not None:
            for i, a in enumerate(_list_pins):
                if not isinstance(a, Pin):
                    raise TypeError(f"_list_pins[{i}] is not a pin configuration (received: {type(a).__name__})")
        
        # Ensure that a valid pandapower network is loaded.
        if not isinstance(self.pp_net, pp.pandapowerNet):
            raise Exception(f"No pandapower network model loaded: {self.pp_net}")

        # Determine model names for worksheet metadata and output file.
        _net_name = "NetworkModel"
        _excel_name = "ePHASORSIM"
        # Prefer data from the pandapower network model if available.
        if self.pp_net.name is not None and self.pp_net.name != "":
            _net_name = self.pp_net.name  
            _excel_name = self.pp_net.name

        # Determine the output path.
        if _excel_file is None or _excel_file == "":
            # If no path is specified, write to the current working directory.
            _cwd = os.getcwd()  # Current working directory.
            _excel_file = f"{_cwd}\\{_excel_name}.xlsx"
        
        # If a directory path is specified, append the generated filename.
        if not _excel_file.endswith(".xlsx"):
            _excel_file = f"{_excel_file}\\{_excel_name}.xlsx"
        self.excel_file = _excel_file
 
        # Copy and open the predefined template.
        _template_path = MODULE_DIR  # Directory of this file.
        shutil.copy(f"{_template_path}\\Template.xlsx", self.excel_file)  
        excel_workbook = load_workbook(self.excel_file)  

        # Create a worksheet for general network metadata.
        excel_workbook_general = excel_workbook.create_sheet("General")
        excel_workbook_general.append(["Excel file version", "v2.0"])               # ePHASORSIM template version.
        excel_workbook_general.append(["Name", _net_name])                          # Network name.
        excel_workbook_general.append(["Frequency (Hz)", self.pp_net.f_hz])         # System frequency.
        excel_workbook_general.append(["Power Base (MVA)", self.pp_net.sn_mva])     # Per-unit base power.
        
        # Create a worksheet for I/O pin mappings.
        excel_worksheet_pins = excel_workbook.create_sheet("Pins")
        if _list_pins is not None:
            # Write one row per pin: type, label, and linked network elements.
            for _oPins in _list_pins:
                excel_worksheet_pins.append(_oPins.to_row())

        # Iterate over all registered network element converter classes.
        for component in self.COMPONENTS:
            # Load all network elements of this type from the network model.
            component_list = component.load_from_pp_net(self.pp_net)
            
            # Only write a worksheet section if components are present.
            if component_list:
                # The component class encapsulates its own worksheet writing logic.
                component.write_excel_worksheet(excel_workbook, component_list)

        # Save the final workbook.
        excel_workbook.save(self.excel_file)

    @classmethod
    def get_components_name_list(cls) -> List[str]:
        """Returns the names of registered network element types with instructions."""
        _lstNames = []
        # Iterate over all registered network element types.
        for component in cls.COMPONENTS:
            # Only include network elements with defined instructions.
            if component.instruction_list != []:
                _lstNames.append(component.__name__)
        return _lstNames