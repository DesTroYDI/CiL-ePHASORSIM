# -*- coding: utf-8 -*-
"""
Base.py
@author: Groß, Hendrik
"""
import re
from abc import ABC, abstractmethod
import pandapower as pp
from typing import List, Any
from openpyxl import Workbook

class BaseComponent(ABC):
    """
    Abstract base class for all network element converters.

    Subclasses must:
        1. Set the class variables (``worksheet_name``, ``model_typ``, ``instruction_list``)
        2. Implement ``to_row()`` and ``load_from_pp_net()``

    Uses the template method pattern: this class defines the Excel writing structure, and subclasses provide the element-specific data.
    """
    
    # ── Class variables (override in subclasses) ─────────────────────
    model_typ: str | None = None        # ePHASORSIM model type string used to find the matching block in the Excel template.
                                        # If set, the writer searches column A starting at row 11 until the string is found and then inserts the network element row after it.
                                        # Reference: https://opal-rt.atlassian.net/wiki/spaces/PEUD/pages/144534926/

    worksheet_name: str | None = None   # Target worksheet name in Excel (e.g. "Bus", "Line", "Load").

    instruction_list: list[str] = []    # Signal/column names for this network element. Elements with an empty list are excluded from the pin system.

    prefix_dict: dict[str, str] = {}    # Maps ID prefix strings to the corresponding pandapower DataFrame name.
                                        # Used by ``clean_and_update_ids()`` to clean component names and write them back into the network model.
                                        # Example: ``{"Ln": "line"}`` -> cleans ``line.name`` and prefixes integers with "Ln0", "Ln1", ...

    # ── Constructor ───────────────────────────────────────────────────────────
    def __init__(self, _id: str | int):
        self.id: str | int = _id

    # ── Abstract methods (must be implemented by subclasses) ────
    @abstractmethod
    def to_row(self) -> List[Any]:
        """Returns the network element data as an ordered list matching the Excel columns. Follow the ePHASORSIM template!"""
        pass

    @classmethod
    @abstractmethod
    def load_from_pp_net(cls, _pp_net: pp.pandapowerNet) -> List[object]:
        """Extracts all instances of this network element type from a pandapower network model."""
        raise NotImplementedError(f"'load_from_pp_net' is not implemented in '{cls}'!")

    # ── Class methods ───────────────────────────────────────────────────────
    @classmethod
    def clean_and_update_ids(cls, _pp_net: pp.pandapowerNet) -> pp.pandapowerNet:
        """
        Cleans network element IDs and writes them back into the pandapower network model.

        For each entry in ``prefix_dict``:
            - Integer or None names receive the configured prefix (e.g. "Ln0").
            - Duplicate IDs are made unique by appending an incrementing counter.
            - Spaces and special characters are removed.

        Notes:
            - Bus IDs are index-based and do not need to be cleaned.
            - VoltageSource IDs are generated in ``load_from_pp_net()`` as "V{idx}".
        """
        if not cls.prefix_dict:
            return _pp_net

        for prefix, dataframe in cls.prefix_dict.items():
            _id_list = []   # List of IDs used to check for duplicates
            if hasattr(_pp_net, dataframe):
                df = getattr(_pp_net, dataframe)

                for idx, row in df.iterrows():
                    _id = cls.clean_id(row["name"], prefix)
                    if _id == prefix:
                        _id = f"{prefix}{idx}"

                    # Resolve duplicate IDs
                    if _id in _id_list:
                        _index = 0
                        while _id in _id_list:
                            _index += 1
                            _id = f"{prefix}{_index}"

                    _id_list.append(_id)
                    df.at[idx, "name"] = _id
        return _pp_net

    @classmethod
    def write_excel_worksheet(cls, excel_workbook: Workbook, component_list):
        """
        Writes all network elements of this class to the matching Excel worksheet.

        If ``model_typ`` is set, the method searches for the type string in column A and inserts rows directly below the corresponding header row.
        Merged cells after the insertion point are shifted down accordingly.
        If ``model_typ`` is "None", writing starts at row 2 (below the column headers).
        """
        if not cls.worksheet_name in excel_workbook.sheetnames:
            raise ValueError(f"Worksheet '{cls.worksheet_name}' not present in the ePHASORSIM template!")
        excel_worksheet = excel_workbook[cls.worksheet_name]

        # Skip rows until ``model_typ`` is found in the first column
        start_row = 2
        if cls.model_typ is not None:
            start_row = 11
            while excel_worksheet.cell(row=start_row, column=1).value != cls.model_typ:
                start_row += 1
                if start_row > 1000:
                    raise Exception(f"Modeltyp '{cls.model_typ}' fehlt in Worksheet '{cls.worksheet_name}'")

            # Skip the header row of this section
            start_row += 2

            # Shift merged cells below the insertion point downward
            _component_count = len(component_list)
            for _merged_cell_range in excel_worksheet.merged_cells.ranges:
                if _merged_cell_range.min_row >= start_row:
                    _merged_cell_range.shift(0, _component_count)

            excel_worksheet.insert_rows(start_row, _component_count)

        # Write network elements
        for component in component_list:
            data_row = component.to_row()
            for col_offset, value in enumerate(data_row):
                col = 1 + col_offset
                excel_worksheet.cell(row=start_row, column=col, value=value)
            start_row += 1

    @classmethod
    def clean_id(cls, _id: str, _prefix: str) -> str:
        """
        Returns a cleaned ID string.

            - None / NaN  -> returns only the prefix (the caller adds the index).
            - Integer     -> prefix + value (e.g. "Ln3").
            - String      -> keeps the content, removes spaces and special characters.
        """
        if _id is not None and not (_id != _id):
            # Try integer conversion (e.g. for float inputs); pure strings remain unchanged.
            try:
                _id = str(int(_id))
            except:
                pass
            
            # If ``_id`` is an integer, prefix it with the configured prefix.
            if str(_id).isdigit():  
                _id = f"{_prefix}{_id}"
            else:
                _id = _id           # Otherwise keep the value as given.
        else:
            # For None or NaN, use only the prefix.
            _id = _prefix

        # Ensure that ``_id`` is a string and remove all spaces.
        _id_tmp = str(_id).replace(" ", "")
        # Remove non-alphanumeric characters.
        return re.sub(r'[^a-zA-Z0-9\s]', '', _id_tmp)