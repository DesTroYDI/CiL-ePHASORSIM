# -*- coding: utf-8 -*-
"""
CiLGenerator.py
@author: Groß, Hendrik
"""
import os, PP2CHIL as PP2CHIL
import dearpygui.dearpygui as dpg
import openpyxl
import pandapower as pp
from .ListEditCtrl import DataGrid, listEditCtrl
from CiL import Controller, Enum, Value

"""
'CiLGenerator' 

The GUI simplifies defining read/write Modbus addresses for an ePHASORSIM Excel template.
"""
class CiLGenerator():
    
    # Global variables
    chil_object: Controller.Controller | None = None                       # Actual CHIL object
    enum_datatypes: list[str] =  Enum.DataType.values()                    # Get data types from enum as a list
    enum_modbus_datatypes: list[str] = Enum.ModbusDataType.values()       # Get Modbus data types from enum as a list
    class_map: dict[str, type] = Controller.Controller.get_class_map()     # Dictionary of all available class objects

    # Constants
    # Static label names that can be mapped automatically to the corresponding pandapower attribute
    MAPPING_KNOWN_LABELS: dict[str, str] = { 
                    "bus_voltage_magnitude": "vm_pu",
                    "bus_voltage_angle": "va_degree",
                    "set_p_gen_load": "set_p_mw",
                    "set_q_gen_load": "set_q_mvar",
                    "profile_set_p_gen_load": "profile_p_mw",
                    "profile_set_q_gen_load": "profile_q_mvar"
                }
    # I/O pin keys ignored in Modbus configuration because they are model inputs
    IGNORE_LABEL: str = "ignore_"

    # DataGrids
    lec_modbus_address = None
    lec_pp_net = None

    def create_gui(self):
        """Creates the interface for generating a CiL configuration."""
        # Popup shown when no dictionary keys have been provided yet
        with dpg.window(id="win_keine_dict_keys", label="Notice",show=False, modal=True, autosize=True):
            dpg.add_text("Error while creating the CHIL Excel configuration. Not all dictionary keys were provided or duplicate values were found!\n\n"
                         "To ensure reliable operation after configuration, dictionary keys must be configured\n"
                         "as specified in the CHIL documentation and must be unique!\n"
                         "Otherwise, equipment values may not be read correctly, e.g.\n"
                         "- internal **Values**:\n"
                            "       - 'p_mw': active power in MW (positive = feed-in)\n"
                            "       - 'q_mvar': reactive power in MVar (positive = capacitive)\n")
            dpg.add_separator()
            dpg.add_button(label="OK", width=75, callback=lambda: dpg.configure_item("win_keine_dict_keys", show=False))

        # File path to the converted pandapower network model
        with dpg.tree_node(id="toolCiLGenerator", label="Generator CHIL Excel Configuration",show=False, default_open=True):
            with dpg.group(horizontal=True, width=0):
                dpg.add_button(label="Save Excel file (*.xlsx)",callback=PP2CHIL.call_file_dialog, user_data=[2,"export_chil_config"])
                dpg.add_input_text(id="export_chil_config", hint="Output path for CHIL Excel configuration (*.xlsx).",width=-1)
            with dpg.group(horizontal=True, width=0):
                dpg.add_button(label="Run: export CHIL Excel configuration", callback=self.create_excel_cfg)
                dpg.bind_item_theme(dpg.last_item(), "theme_execute")
                dpg.add_loading_indicator(tag="loadinc_export_chil_config", show=False)
                dpg.add_text(tag="status_export_chil_config", default_value="No export started yet", color=(255, 0, 0))
            with dpg.tab_bar(id="tab_control_chil"):
                pass
                
        # Popup shown when no file has been selected yet
        with dpg.window(id="win_kein_ephasorsim_net", label="Notice",show=False, modal=True, autosize=True):
            dpg.add_text("No ePHASORSIM model (Excel file) has been selected yet, or an error occurred while loading the ePHASORSIM model!")
            dpg.add_separator()
            dpg.add_button(label="OK", width=75, callback=lambda: dpg.configure_item("win_kein_ephasorsim_net", show=False))

    def load_ephasorsim_chil(self,_excel_file, _pickle_file) -> tuple[bool, str]:
        """Loads an ePHASORSIM model."""
        # Initialize variables
        str_return: str = ""
        bool_return: bool = False
        try:
            # Load Excel file and select worksheet 'Pins'
            wb = openpyxl.load_workbook(_excel_file, read_only=True)
            self.ws = wb["Pins"]

            # Load pandapower network model
            _pp_net = pp.from_pickle(_pickle_file, True)

            # DataGrid for Modbus component input
            dpg.delete_item("tab_modbus_adress", children_only=False)
            dpg.delete_item("tab_pp_reference", children_only=False)

            # Modbus address configuration
            with dpg.tab(id="tab_modbus_adress",label="Modbus Address Configuration", parent="tab_control_chil"):
                dpg.add_text("[*] - These columns are informational only and are not used in calculations.")
                dpg.add_text("[Dictionary key] - Specifies the corresponding pandapower attribute (e.g., Generator -> 'p_mw').")
                dpg.add_text("[Register/bit count] - Default is float32. Specify the number of registers to use; for 'Coil' and 'Discrete Input' this value is automatically 1 and can be ignored. This value is considered when creating the Excel configuration.")
                dpg.add_text("[Modbus start address] - Start address of the first object in pin configuration. Additional objects are addressed sequentially during Excel configuration creation based on the configured object count.")
                dpg.add_text(f"** In this view, all I/O pins with label '{self.IGNORE_LABEL}*' are ignored, because they are assumed to be model inputs defined by a profile")
                # Set default IDs for comboboxes of Modbus data types and data types
                idx_modbus_datatype = self.enum_modbus_datatypes.index("input_register")
                idx_datatype = self.enum_datatypes.index("float32")
                
                # Remove existing ListEditControl to load a new configuration
                dg_io_pins = DataGrid(
                    title="Modbus Address Configuration",     
                    columns = ["Object Count [*]","Instruction Key [*]","I/O Label [*]", "Dictionary Key","Modbus Data Type","Data Type","Modbus Start Address", "Scaling", "Unit"],   
                    dtypes = [DataGrid.TXT_TEXT,DataGrid.TXT_TEXT,DataGrid.TXT_TEXT, DataGrid.TXT_STRING,DataGrid.COMBO,DataGrid.COMBO,DataGrid.TXT_INT, DataGrid.TXT_FLOAT, DataGrid.TXT_STRING],
                    defaults=["-","-","","",idx_modbus_datatype,idx_datatype,-1,1.0,""],
                    combo_lists=[None,None,None,None,self.enum_modbus_datatypes,self.enum_datatypes,None,None,None]
                )

                # Create one Modbus address configuration from top to bottom for each input/output signal
                for pin_row in self.ws.iter_rows():
                    # Determine I/O label
                    label = str(pin_row[1].value)

                    # Ignore these I/O pins (model inputs)
                    if label.startswith(self.IGNORE_LABEL):
                        continue

                    # Automatically set known bus labels
                    dict_key = ""
                    if label in self.MAPPING_KNOWN_LABELS:
                        dict_key = self.MAPPING_KNOWN_LABELS[label]

                    # Determine instruction
                    first_instruction = str(pin_row[2].value)
                    instruction = first_instruction[first_instruction.find("/") + 1:]

                    # Determine number of objects
                    count_elements = max((cell.column for cell in pin_row if cell.value is not None),default=0) - 2

                    # Insert data row (default values with defined modifications)
                    default_row = dg_io_pins.defaults.copy()
                    default_row[0] = count_elements
                    default_row[1] = instruction
                    default_row[2] = label
                    default_row[3] = dict_key
                    
                    # If INCOMING pin, choose "Holding Register" as default Modbus data type
                    if str(pin_row[0].value).upper() == "INCOMING":
                        default_row[4] = self.enum_modbus_datatypes.index("holding_register")

                    dg_io_pins.append(default_row)
                
                # Determine start addresses based on existing Modbus data types
                # Determine column indices
                idx_obj_count = dg_io_pins.columns.index("Object Count [*]")
                idx_start_address = dg_io_pins.columns.index("Modbus Start Address")
                idx_modbus_datatyp = dg_io_pins.columns.index("Modbus Data Type")
                # Create dictionary from existing Modbus data types and initialize with 0
                modbus_datatyp_list = set(dg_io_pins.data[idx_modbus_datatyp])
                start_addresses = dict.fromkeys(modbus_datatyp_list, 0)
                # Iterate through all data rows
                for idx in range(len(dg_io_pins.data[0])):
                    # Update start address in grid
                    dg_io_pins.data[idx_start_address][idx] = start_addresses[dg_io_pins.data[idx_modbus_datatyp][idx]]
                    # Get number of objects
                    count_obj = dg_io_pins.data[idx_obj_count][idx]
                    # Increment start address by the number of objects
                    start_addresses[dg_io_pins.data[idx_modbus_datatyp][idx]] += (2 * count_obj)

                # Create new ListEditControl
                self.lec_modbus_address,_= listEditCtrl(dpg.generate_uuid(),grid=dg_io_pins, allow_add=False, allow_movement=False)

            # Load pandapower network model and map references
            with dpg.tab(id="tab_pp_reference", label="Pandapower Element Mapping", parent="tab_control_chil"):
                dpg.add_text("[*] - These columns are informational only and are not used in calculations (read-only).")
                dpg.add_text("Allows assigning objects to a component class and referencing a pandapower network element. Initially, an automatic assignment is attempted based on object ID (net index).\n\n"
                            "Additionally, you can define whether elements should be included in Modbus address configuration")
                # Determine pandapower DataFrames
                class_keys = list(self.class_map.keys())
                pp_dfs = []
                for cls in self.class_map.values():
                    if cls.df_pp is not None and cls.df_pp != "":
                        pp_dfs.append(cls.df_pp)
                
                # Create DataGrid
                dg_objects = DataGrid(
                    title="Pandapower Network Model Mapping",     
                    columns = ["Object ID [*]","Network Element", "PandaPower-NetIndex","PandaPower-DataFrame [*]", "Object Information [*]"],
                    dtypes = [DataGrid.TXT_TEXT,DataGrid.COMBO,DataGrid.TXT_INT,DataGrid.TXT_TEXT,DataGrid.TXT_TEXT],
                    defaults=["",0,-1,"",""],
                    combo_lists=[None,class_keys,None,pp_dfs,None]
                )

                # Determine all objects
                dict_objects = {}
                for pin_row in self.ws.iter_rows():
                    count_elements = max((cell.column for cell in pin_row if cell.value is not None),default=0)

                    # Iterate over all objects and evaluate objects starting from the second column
                    for column_index in range(count_elements):
                        if column_index > 1:
                            # Remove instruction from cell value
                            cell_value = self.ws.cell(row=pin_row[0].row, column=(column_index+1)).value  # type: ignore
                            obj_id = cell_value[:cell_value.find("/")]

                            # Add object if not already present
                            if obj_id not in dict_objects:
                                # Try to determine the network element class automatically from the name
                                # !! If the name cannot be mapped to a class, assume it is a 'Bus'
                                # !! Reason: 'Bus' elements are the only ones directly identified by DataFrame index
                                obj_net_index = -1
                                obj_class = list(self.class_map.keys())[0]
                                # Check which class this object belongs to
                                for k,v in self.class_map.items():
                                    try:
                                        # Determine/check object class and net index
                                        index = pp.get_element_index(_pp_net, v.df_pp, obj_id, True)
                                        obj_class = k
                                        obj_net_index = int(index)
                                        break
                                    except:
                                        # Bus elements use net index directly as ID. These are taken over directly
                                        if str(obj_id).isdigit():
                                            obj_net_index = int(obj_id)
                                            break
                                # Store result
                                dict_objects[obj_id] = [obj_class, obj_net_index]

                # Column IDs
                idx_objid = dg_objects.columns.index("Object ID [*]")
                idx_cls_name = dg_objects.columns.index("Network Element")
                idx_net_index = dg_objects.columns.index("PandaPower-NetIndex")

                # Insert all objects into the table
                class_key_list = list(self.class_map.keys())
                for obj_id in dict_objects:
                    # Get result data
                    dict_result = dict_objects[obj_id]

                    class_name = class_key_list.index(dict_result[0])
                    # Take default row and apply adjustments
                    default_row = dg_objects.defaults.copy()
                    default_row[idx_objid]=obj_id
                    default_row[idx_cls_name]=class_name
                    default_row[idx_net_index]=dict_result[1]

                    dg_objects.append(default_row)
                
                # Generate object list
                self.lec_pp_net,ref_pp= listEditCtrl(dpg.generate_uuid(),grid=dg_objects,allow_add=False, allow_delete=False, allow_movement=False, use_filter=True)
                ref_pp(_pp_net)

            # Excel file loaded successfully
            str_return = "ePHASORSIM loaded successfully"
            bool_return = True 
        except Exception as e:
            str_return = f"Error while loading Excel file: {e}"
        
        return (bool_return,str_return)

    def create_excel_cfg(self,sender, app_data, user_data):
        """Creates an Excel configuration from the defined settings."""
        dpg.configure_item("loadinc_export_chil_config", show=True) 
        dpg.configure_item("status_export_chil_config", default_value="Running...")

        # 1. Check whether configuration steps are already completed
        if self.lec_modbus_address is None or self.lec_pp_net is None:
            dpg.configure_item("loadinc_export_chil_config", show=False) 
            dpg.configure_item("status_export_chil_config", default_value="Configuration not completed yet!")
            dpg.configure_item("status_export_chil_config", color=(255, 0, 0))
            return

        # 2. Check whether an Excel file is specified; otherwise use current folder and name it "CiL-Configuration.xlsx"
        file_path = dpg.get_value("export_chil_config")
        if file_path is None or file_path == "":
            # No path was provided
            file_path = f"{os.getcwd()}\\CiL-Configuration.xlsx"
        elif not file_path.endswith(".xlsx"):
            # A folder was provided
            file_path = f"{file_path}\\CiL-Configuration.xlsx"
        
        # Retrieve DataGrid data
        grid_modbus_address = self.lec_modbus_address()
        grid_pp_net = self.lec_pp_net()

        # 3. Check whether all dictionary keys are specified and unique
        idx_dict_key = grid_modbus_address.columns.index("Dictionary Key")
        keys_list = grid_modbus_address.data[idx_dict_key]
        dict_keys_empty = ("" in keys_list) or (None in keys_list)
        dict_keys_unique = len(keys_list) != len(set(keys_list))
        if dict_keys_empty or dict_keys_unique:
            dpg.configure_item("win_keine_dict_keys", show=True)
            PP2CHIL.center_window("win_keine_dict_keys")
            dpg.configure_item("loadinc_export_chil_config", show=False) 
            if dict_keys_empty:
                dpg.configure_item("status_export_chil_config", default_value="No dictionary keys specified!")
            else:
                dpg.configure_item("status_export_chil_config", default_value="Duplicate dictionary keys specified!")
            dpg.configure_item("status_export_chil_config", color=(255, 0, 0))
            return
        
        # Determine column indices from headers
        idx_io_bezeichnung =grid_modbus_address.columns.index("I/O Label [*]") 
        idx_address = grid_modbus_address.columns.index("Modbus Start Address")
        idx_scale = grid_modbus_address.columns.index("Scaling")
        idx_unit = grid_modbus_address.columns.index("Unit")
        idx_modbus_data_typ = grid_modbus_address.columns.index("Modbus Data Type")
        idx_data_typ = grid_modbus_address.columns.index("Data Type")

        # Run export
        try:
            # Iterate through Excel file again and enrich with additional information from ListEditControl
            # Convert data into IModbusElements in the end
            dict_components = {}
            row_idx = -1        
            for pin_row in self.ws.iter_rows():
                # Consider I/O pin row only if it exists in address configuration
                if not pin_row[1].value in grid_modbus_address.data[idx_io_bezeichnung]:
                    continue
                # Row indices
                row_idx += 1                    # Row index of Modbus address configuration
                row_idx_ws = pin_row[0].row     # Row index of Excel worksheet

                # Number of objects
                count_elements = max((cell.column for cell in pin_row if cell.value is not None),default=0)

                # Store values that apply to the entire column (Modbus components, ModbusValue)
                dict_key = grid_modbus_address.data[idx_dict_key][row_idx]
                address = grid_modbus_address.data[idx_address][row_idx]
                scale = grid_modbus_address.data[idx_scale][row_idx]
                unit = grid_modbus_address.data[idx_unit][row_idx] if grid_modbus_address.data[idx_unit][row_idx] != "" else None

                # Calculate address increment and obtain data type from selected data type
                name_data_typ = self.enum_datatypes[int(grid_modbus_address.data[idx_data_typ][row_idx])] 
                name_modbus_data_typ = self.enum_modbus_datatypes[int(grid_modbus_address.data[idx_modbus_data_typ][row_idx])] 
                data_typ = Enum.DataType.from_label(name_data_typ)
                modbus_data_typ = Enum.ModbusDataType(name_modbus_data_typ)

                # Calculate address increment
                if modbus_data_typ.is_bit_type:
                    addr_inc = 1
                else:
                    addr_inc = data_typ.register_count

                # Iterate over all objects and evaluate from the second column
                for column_idx in range(count_elements):
                    if column_idx > 1:
                        # Determine column indices from headers
                        idx_obj_id = grid_pp_net.columns.index("Object ID [*]")
                        idx_netzelement = grid_pp_net.columns.index("Network Element")
                        idx_net_index = grid_pp_net.columns.index("PandaPower-NetIndex")

                        # Remove instruction from cell value
                        cell_value = self.ws.cell(row=(row_idx_ws), column=(column_idx+1)).value # type: ignore
                        obj_id = cell_value[:cell_value.find("/")]

                        # Determine grid index in pandapower ListEditControl
                        grid_idx = grid_pp_net.data[idx_obj_id].index(obj_id)

                        # Create new object if no component exists yet
                        if obj_id not in dict_components:
                            # Read class description from DataGrid
                            class_idx = int(grid_pp_net.data[idx_netzelement][grid_idx])
                            class_name = list(self.class_map.keys())[class_idx]
                            if class_name not in self.class_map:
                                raise KeyError(f"Class instance for '{class_name}' cannot be instantiated!")
                            cls = self.class_map[class_name]
                            # Get net index from DataGrid
                            net_index = grid_pp_net.data[idx_net_index][grid_idx]
                            # Instantiate new class instance and store it in dictionary
                            dict_components[obj_id] = cls(obj_id,net_index, {})
                        
                        # Add ModbusValue object
                        obj_modbus_value = Value.ModbusValue(modbus_data_typ,data_typ,address,unit,scale)
                        obj = dict_components[obj_id]
                        obj.values[dict_key]= obj_modbus_value

                        # Increment address value
                        address += addr_inc
                            
            # Execute export
            list_components = list(dict_components.values())
            Controller.Controller.write_cfg_to_excel(list_components,file_path)

            dpg.configure_item("status_export_chil_config", default_value="Export successful!")
            dpg.configure_item("status_export_chil_config", color=(0, 255, 0))
        except Exception as e:
            dpg.configure_item("status_export_chil_config", default_value=str(e))
            dpg.configure_item("status_export_chil_config", color=(255, 0, 0))
        dpg.configure_item("loadinc_export_chil_config", show=False) 
       